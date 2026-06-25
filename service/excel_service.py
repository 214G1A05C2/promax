import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
REPORT_DIR = "reports"

class ExcelService:
    def __init__(self, output_dir: str = REPORT_DIR):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_intent_report(self, data: List[Dict[str, Any]]):
        wb = self._create_workbook("Intent Analysis")
        ws = wb.active
        headers = [
            "Call Id",
            "Date",
            "Clinic Name",
            "Call Transcript",
            "Primary Intent",
            "Secondary Intents",
            # "Confidence",
            # "Needs Human Review",
            # "Reasoning",
        ]
        self._write_header(ws, headers)
        
        for row in data:
            date_value = row.get("Date")

            if isinstance(date_value, datetime) and date_value.tzinfo is not None:
                date_value = date_value.replace(tzinfo=None)

            primary_display, secondary_str = self._format_intent_columns(
                row.get("primary_intent"),
                row.get("secondary_intents"),
            )

            transcript_display = self._format_transcript(
                row.get("call_transcript")
            )

            ws.append([
                row["Call ID"],
                date_value,
                row["clinic_name"],
                transcript_display,
                primary_display,
                secondary_str,
            ])
        self._format_sheet(ws, len(headers))
        self._save(wb, "Intent_Analysis_Report.xlsx")

    def _format_transcript(self, transcript: Any) -> str:
        if transcript is None:
            return ""
        if isinstance(transcript, (dict, list)):
            return json.dumps(transcript, ensure_ascii=False, indent=2)
        return str(transcript)

    def _format_intent_columns(self, primary_intent: Any, secondary_intents: Any) -> (str, str):
        primary = "" if primary_intent is None else str(primary_intent).strip()
        if not primary:
            return "", ""

        secondaries: List[str] = []
        if isinstance(secondary_intents, list):
            secondaries = [str(i).strip() for i in secondary_intents if str(i).strip()]
        elif isinstance(secondary_intents, str) and secondary_intents.strip():
            # Handles rows where secondary intents are already flattened text.
            secondaries = [part.strip() for part in secondary_intents.split(",") if part.strip()]

        unique_secondaries: List[str] = []
        for intent in secondaries:
            if intent != primary and intent not in unique_secondaries:
                unique_secondaries.append(intent)

        primary_parts = [primary]
        if unique_secondaries:
            primary_parts.append(unique_secondaries[0])

        primary_display = " / ".join(primary_parts)

        # Secondary column should only show intents when total intents exceed two.
        remaining = unique_secondaries[1:] if len(primary_parts) == 2 else unique_secondaries
        secondary_display = ", ".join(remaining) if remaining else ""

        return primary_display, secondary_display

    def _create_workbook(self, sheet_title: str) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.append([f"{sheet_title} Report"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
        ts_cell = ws.cell(row=2, column=1)
        ts_cell.font = Font(italic=True, size=10)
        ts_cell.alignment = Alignment(horizontal="center")
        ws.append([])
        return wb

    def _write_header(self, ws, headers: List[str], start_row: int = 4):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

    def _format_sheet(self, ws, num_columns: int):
        header_row = 4
        data_start = header_row + 1
        max_row = ws.max_row
        if max_row > header_row:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_columns)}{max_row}"
            ws.freeze_panes = f"A{data_start}"
        for col in range(1, num_columns + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(header_row, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[col_letter].width = adjusted_width
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=num_columns):
            for cell in row:
                cell.border = thin_border
                if cell.row > header_row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _save(self, wb: Workbook, filename: str):
        path = os.path.join(self.output_dir, filename)
        wb.save(path)
        logger.info(f"Report saved: {path}")
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        """
        for row in data:
            primary_display, secondary_str = self._format_intent_columns(
                row.get("primary_intent"),
                row.get("secondary_intents"),
            )
            transcript_display = self._format_transcript(row.get("call_transcript"))
            ws.append([
                row["Call ID"],
                row["Date"],
                row["clinic_name"],
                transcript_display,
                primary_display,
                secondary_str,
                # row.get("confidence", 0),
                # row.get("needs_human_review", False),
                # row.get("reasoning", ""),
            ])
        """